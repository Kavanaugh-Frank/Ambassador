import openpyxl
import csv
from datetime import datetime
from flask import Flask, request, jsonify, render_template
import os

import csv
from datetime import datetime

EXCEL_FILE = "Event.xlsx"
CSV_FILE = "Output.csv"

class AmbassadorOfTheMonthProcessor:
    def __init__(self, path, csv_file_path, selected_month, events_factor=1):
        self.path = path
        self.csv_file_path = csv_file_path
        self.events_factor = events_factor
        self.selected_month = selected_month
        self.total_time = {}
    def process_events(self):
        # Prepare the CSV file for output
        with open(self.csv_file_path, mode='w', newline='', encoding='utf-8') as csvfile:
            csvwriter = csv.writer(csvfile)
            
            # Write the header row
            csvwriter.writerow(['Event Name', 'Time Difference', 'People There'])
            
            # Iterate through all the rows in the worksheet
            for row_num in range(3, self.sheet.max_row + 1):
                # Check if the cell value is a datetime object
                cell_value = self.sheet.cell(row=row_num, column=1).value
                event_month = None
                
                if isinstance(cell_value, datetime):
                    event_month = cell_value.month
                else:
                    continue  # Skip rows where the cell is not a datetime object

                if event_month != int(self.selected_month):
                    continue  # Skip rows where the month does not match

                end_time = self.sheet.cell(row=row_num, column=3).value
                start_time = self.sheet.cell(row=row_num, column=2).value

                if end_time is not None and end_time != "TBD" and start_time is not None and start_time != "TBD":
                    # Convert `datetime.time` to `datetime.datetime` for the same arbitrary date
                    arbitrary_date = datetime(2000, 1, 1)  # Or any arbitrary date

                    # Combine the arbitrary date with the time objects to get datetime objects
                    end_datetime = datetime.combine(arbitrary_date, end_time)
                    start_datetime = datetime.combine(arbitrary_date, start_time)

                    # Calculate the difference between the two datetime objects
                    time_diff = (end_datetime - start_datetime) * self.events_factor

                    # Convert the time difference to minutes
                    time_diff_minutes = float(time_diff.total_seconds() / 60)

                    # Prepare event name and people list
                    event_name = self.sheet.cell(row=row_num, column=4).value
                    people_list = [self.sheet.cell(row=1, column=col_num).value 
                                    for col_num in range(8, self.sheet.max_column + 1) 
                                    if self.sheet.cell(row=row_num, column=col_num).value in ["x", "X"]]

                    people = ', '.join(people_list)

                    # Write the row to the CSV file with time difference in minutes
                    csvwriter.writerow([event_name, time_diff_minutes, people])
    def process_tours(self):
        with open(self.csv_file_path, mode='a', newline='', encoding='utf-8') as csvfile:
            csvwriter = csv.writer(csvfile)
            
            # Iterate through all the rows in the worksheet
            for row_num in range(3, self.sheet.max_row + 1):
                cell_value = self.sheet.cell(row=row_num, column=1).value
                event_month = None
                
                if isinstance(cell_value, datetime):
                    event_month = cell_value.month
                else:
                    continue  # Skip rows where the cell is not a datetime object

                if event_month != int(self.selected_month):
                    continue  # Skip rows where the month does not match

                end_time = self.sheet.cell(row=row_num, column=3).value
                start_time = self.sheet.cell(row=row_num, column=2).value

                if end_time is not None and end_time != "TBD" and start_time is not None and start_time != "TBD":
                    # Convert `datetime.time` to `datetime.datetime` for the same arbitrary date
                    arbitrary_date = datetime(2000, 1, 1)  # Or any arbitrary date

                    # Combine the arbitrary date with the time objects to get datetime objects
                    end_datetime = datetime.combine(arbitrary_date, end_time)
                    start_datetime = datetime.combine(arbitrary_date, start_time)

                    # Calculate the difference between the two datetime objects
                    time_diff = (end_datetime - start_datetime) * self.events_factor

                    # Convert the time difference to minutes
                    time_diff_minutes = float(time_diff.total_seconds() / 60)

                    # Prepare event name and people list
                    event_name = "Tour"
                    people_list = [self.sheet.cell(row=1, column=col_num).value 
                                    for col_num in range(7, self.sheet.max_column + 1) 
                                    if self.sheet.cell(row=row_num, column=col_num).value in ["x", "X"]]

                    people = ', '.join(people_list)

                    # Write the row to the CSV file with time difference in minutes
                    csvwriter.writerow([event_name, time_diff_minutes, people])
    # goes through the CSV and adds the total time spent by each person
    def calculate_total_time(self):
        with open(self.csv_file_path, mode='r', newline='', encoding='utf-8') as csvfile:
            csv_reader = csv.reader(csvfile)
            next(csv_reader)  # Skip the header row
            
            for row in csv_reader:
                for person in row[2].split(","):
                    person = person.replace(" ", "")
                    if person:
                        if person not in self.total_time:
                            self.total_time[person] = float(row[1])
                        else:
                            self.total_time[person] += float(row[1])
    # loads the correct workbook for the process functions
    def load_workbook(self, workbook):
        # Load the Excel workbook and select the worksheet
        self.wb = openpyxl.load_workbook(self.path)
        self.sheet = self.wb[workbook]
    def run(self):
        self.load_workbook("EVENTS")
        self.process_events()
        self.load_workbook("TOURS")
        self.process_tours()
        self.calculate_total_time()
        print(f"Data has been written to {self.csv_file_path}")
        print(self.total_time)
        return self.total_time
    def num_events(self):
        self.load_workbook("EVENTS")
        people = {}
        with open(self.csv_file_path, mode='r', newline='', encoding='utf-8') as csvfile:
            for row in range(3, self.sheet.max_row + 1):
                for col in range(8, self.sheet.max_column + 1):
                    if self.sheet.cell(column=col, row=row).value == 'x' or self.sheet.cell(column=col, row=row).value == 'X':
                        name = self.sheet.cell(row=1, column=col).value
                        if name not in people:
                            people[name] = 1
                        else:
                            people[name] += 1
        return people

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('./index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files.get('file')
    month = request.form.get('month')

    # hard coded ATM
    file.save(EXCEL_FILE)

    processor = AmbassadorOfTheMonthProcessor(EXCEL_FILE, CSV_FILE, month)
    result = processor.run()
    # print(processor.num_events())
    num_events = processor.num_events()
    sorted_result = sorted(result.items(), key=lambda x:x[1], reverse=True)
    return jsonify({
        "sorted_result": sorted_result,
        "num_events": num_events
    })


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8080, debug=True)
